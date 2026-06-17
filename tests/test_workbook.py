from __future__ import annotations

from openpyxl import load_workbook

from intrastat_generator.cn import CnResolver
from intrastat_generator.models import IntrastatItem, TariffEntry
from intrastat_generator.workbook import WorkbookBuilder


def test_workbook_builder_writes_settings_and_route_cost_sheets(tmp_path):
    item = IntrastatItem(
        poz_id="1",
        opis="SZAFKA RTV",
        country="DE",
        delivery_terms="",
        transaction_type="11",
        source_cn="",
        transport_type="",
        origin_country="PL",
        mass_net=10,
        supplementary_qty="",
        invoice_value=1000,
        statistical_value="",
        vat_id="DE123",
        attrs={},
    )
    tariff_entries = [
        TariffEntry(
            code="94036010",
            spaced_code="9403 60 10",
            description="Meble drewniane",
            path_text="Taryfa testowa",
        )
    ]
    route_config = {
        "origin_voivodeship": "podkarpackie",
        "allocation_basis": "invoice_value",
        "use_invoice_cap": True,
        "max_transport_share_pct": 8.0,
        "routes": [
            {
                "active": True,
                "country": "DE",
                "zone": "STANDARD",
                "foreign_cost_pln": 200,
                "truck_count": 1,
                "max_correction_pct": 8.0,
                "default": True,
                "note": "test",
            }
        ],
    }
    config = {
        "cn_confident_threshold": 90.0,
        "cn_uncertain_threshold": 80.0,
        "statistical_value_mode": "blank",
    }
    builder = WorkbookBuilder(
        dicts={},
        tariff_entries=tariff_entries,
        resolver=CnResolver(tariff_entries, 90.0, 80.0),
        config=config,
        route_config=route_config,
    )
    output_path = tmp_path / "intrastat.xlsx"

    builder.build([item], {"Typ": "W"}, output_path)

    workbook = load_workbook(output_path, read_only=True)
    try:
        assert "Ustawienia" in workbook.sheetnames
        assert "Koszty_transportu" in workbook.sheetnames
    finally:
        workbook.close()
