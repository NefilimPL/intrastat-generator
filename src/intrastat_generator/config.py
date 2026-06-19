from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict

from openpyxl.styles import Border, Font, PatternFill, Side

APP_NAME_BASE = "Generator INTRASTAT XLSX"


def app_name(version: str) -> str:
    return f"{APP_NAME_BASE} {version}"

CONFIG_FILE = "config.json"
CONFIG_DIR_NAME = "Intrastat generator config"
OUTPUT_DIR_NAME = "wygenerowane_xlsx"
DICT_DIR_NAME = "slowniki"
LOG_DIR_NAME = "logi"

ROUTE_COSTS_FILE = "koszty_transportu.json"

VOIVODESHIPS = [
    "dolnośląskie", "kujawsko-pomorskie", "lubelskie", "lubuskie", "łódzkie", "małopolskie",
    "mazowieckie", "opolskie", "podkarpackie", "podlaskie", "pomorskie", "śląskie",
    "świętokrzyskie", "warmińsko-mazurskie", "wielkopolskie", "zachodniopomorskie",
]

# Domyślne kwoty są roboczą tabelą startową do edycji w GUI.
# Oznaczają SZACOWANY koszt zagranicznego odcinka dla jednego pełnego TIR-a,
# czyli koszt od granicy Polski do kraju/strefy docelowej. Dla większych krajów
# warto rozdzielić trasę na część kraju, np. DE_EAST/DE_WEST.
DEFAULT_ROUTE_COSTS: Dict[str, Any] = {
    "origin_voivodeship": "podkarpackie",
    "allocation_basis": "invoice_value",  # mass_net | invoice_value
    "use_invoice_cap": True,
    "max_transport_share_pct": 8.0,
    "routes": [
        {"active": True, "country": "AT", "zone": "STANDARD", "foreign_cost_pln": 3500, "truck_count": 1, "default": True, "note": "Austria - domyślna robocza stawka"},
        {"active": True, "country": "BE", "zone": "STANDARD", "foreign_cost_pln": 6500, "truck_count": 1, "default": True, "note": "Belgia - przez DE"},
        {"active": True, "country": "BG", "zone": "STANDARD", "foreign_cost_pln": 5200, "truck_count": 1, "default": True, "note": "Bułgaria"},
        {"active": True, "country": "CY", "zone": "STANDARD", "foreign_cost_pln": 12000, "truck_count": 1, "default": True, "note": "Cypr - wymaga ręcznej weryfikacji promu"},
        {"active": True, "country": "CZ", "zone": "STANDARD", "foreign_cost_pln": 1200, "truck_count": 1, "default": True, "note": "Czechy"},
        {"active": True, "country": "DE", "zone": "EAST", "foreign_cost_pln": 2200, "truck_count": 1, "default": False, "note": "Niemcy wschód"},
        {"active": True, "country": "DE", "zone": "CENTER", "foreign_cost_pln": 3800, "truck_count": 1, "default": False, "note": "Niemcy centrum"},
        {"active": True, "country": "DE", "zone": "WEST", "foreign_cost_pln": 5200, "truck_count": 1, "default": True, "note": "Niemcy zachód - domyślna strefa"},
        {"active": True, "country": "DE", "zone": "NORTH", "foreign_cost_pln": 4500, "truck_count": 1, "default": False, "note": "Niemcy północ"},
        {"active": True, "country": "DE", "zone": "SOUTH", "foreign_cost_pln": 4800, "truck_count": 1, "default": False, "note": "Niemcy południe"},
        {"active": True, "country": "DK", "zone": "STANDARD", "foreign_cost_pln": 5600, "truck_count": 1, "default": True, "note": "Dania"},
        {"active": True, "country": "EE", "zone": "STANDARD", "foreign_cost_pln": 4500, "truck_count": 1, "default": True, "note": "Estonia"},
        {"active": True, "country": "ES", "zone": "STANDARD", "foreign_cost_pln": 12500, "truck_count": 1, "default": True, "note": "Hiszpania"},
        {"active": True, "country": "FI", "zone": "STANDARD", "foreign_cost_pln": 9000, "truck_count": 1, "default": True, "note": "Finlandia"},
        {"active": True, "country": "FR", "zone": "EAST", "foreign_cost_pln": 8500, "truck_count": 1, "default": True, "note": "Francja wschód"},
        {"active": True, "country": "FR", "zone": "WEST", "foreign_cost_pln": 11000, "truck_count": 1, "default": False, "note": "Francja zachód"},
        {"active": True, "country": "FR", "zone": "SOUTH", "foreign_cost_pln": 12500, "truck_count": 1, "default": False, "note": "Francja południe"},
        {"active": True, "country": "GR", "zone": "STANDARD", "foreign_cost_pln": 11000, "truck_count": 1, "default": True, "note": "Grecja"},
        {"active": True, "country": "HR", "zone": "STANDARD", "foreign_cost_pln": 5500, "truck_count": 1, "default": True, "note": "Chorwacja"},
        {"active": True, "country": "HU", "zone": "STANDARD", "foreign_cost_pln": 3500, "truck_count": 1, "default": True, "note": "Węgry"},
        {"active": True, "country": "IE", "zone": "STANDARD", "foreign_cost_pln": 13000, "truck_count": 1, "default": True, "note": "Irlandia - prom, do weryfikacji"},
        {"active": True, "country": "IT", "zone": "NORTH", "foreign_cost_pln": 8000, "truck_count": 1, "default": True, "note": "Włochy północ"},
        {"active": True, "country": "IT", "zone": "CENTER", "foreign_cost_pln": 10000, "truck_count": 1, "default": False, "note": "Włochy centrum"},
        {"active": True, "country": "IT", "zone": "SOUTH", "foreign_cost_pln": 12500, "truck_count": 1, "default": False, "note": "Włochy południe"},
        {"active": True, "country": "LT", "zone": "STANDARD", "foreign_cost_pln": 3000, "truck_count": 1, "default": True, "note": "Litwa"},
        {"active": True, "country": "LU", "zone": "STANDARD", "foreign_cost_pln": 6200, "truck_count": 1, "default": True, "note": "Luksemburg"},
        {"active": True, "country": "LV", "zone": "STANDARD", "foreign_cost_pln": 4000, "truck_count": 1, "default": True, "note": "Łotwa"},
        {"active": True, "country": "MT", "zone": "STANDARD", "foreign_cost_pln": 14000, "truck_count": 1, "default": True, "note": "Malta - prom, do weryfikacji"},
        {"active": True, "country": "NL", "zone": "STANDARD", "foreign_cost_pln": 6200, "truck_count": 1, "default": True, "note": "Holandia - przez DE"},
        {"active": True, "country": "PT", "zone": "STANDARD", "foreign_cost_pln": 15000, "truck_count": 1, "default": True, "note": "Portugalia"},
        {"active": True, "country": "RO", "zone": "STANDARD", "foreign_cost_pln": 5200, "truck_count": 1, "default": True, "note": "Rumunia"},
        {"active": True, "country": "SE", "zone": "STANDARD", "foreign_cost_pln": 8500, "truck_count": 1, "default": True, "note": "Szwecja"},
        {"active": True, "country": "SI", "zone": "STANDARD", "foreign_cost_pln": 4300, "truck_count": 1, "default": True, "note": "Słowenia"},
        {"active": True, "country": "SK", "zone": "STANDARD", "foreign_cost_pln": 1500, "truck_count": 1, "default": True, "note": "Słowacja"},
    ],
}


FORBIDDEN_DESC_CHARS = ["&", '"', "'", "<", ">", ";"]

OUTPUT_COLUMNS = [
    "Opis towaru",                 # A - tekst
    "Kod kraju",                  # B - tekst
    "Warunki dostawy",            # C - tekst, opcjonalne
    "Rodzaj transakcji",          # D - liczba wg instrukcji, ale zapis tekstowy pomaga nie gubić zer
    "Kod towaru CN",              # E - liczba wg instrukcji, zapis tekstowy chroni format 8 cyfr
    "Rodzaj transportu",          # F - liczba wg instrukcji, opcjonalne
    "Kraj pochodzenia",           # G - tekst
    "Masa netto kg",              # H - liczba całkowita
    "Ilość w jedn. uzup.",        # I - liczba całkowita/opcjonalna
    "Wartość fakturowa PLN",      # J - liczba całkowita
    "Wartość statystyczna PLN",   # K - liczba całkowita/opcjonalna
    "VAT kontrahenta",            # L - tekst
]

DEFAULT_CONFIG: Dict[str, Any] = {
    "tariff_path": "",
    "tariff_year": "",
    "dict_dir": DICT_DIR_NAME,
    "output_dir": OUTPUT_DIR_NAME,
    "default_delivery_terms": "",
    "default_transaction_type": "11",
    "default_transport_type": "",
    "statistical_value_mode": "blank",  # blank | copy_invoice_when_required | copy_invoice_always | subtract_foreign_transport_by_route
    "origin_voivodeship": "podkarpackie",
    "transport_allocation_basis": "invoice_value",  # mass_net | invoice_value
    "transport_costs_file": ROUTE_COSTS_FILE,
    "transport_use_invoice_cap": True,
    "transport_max_share_pct": 8.0,
    "auto_open_output_folder": False,
    "cn_confident_threshold": 90.0,
    "cn_uncertain_threshold": 80.0,
    "hide_dictionary_sheets": False,
}

STATUS_OK = "OK"
STATUS_UNCERTAIN = "NIEPEWNY"
STATUS_MISSING = "BRAK"

FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
FILL_SUBHEADER = PatternFill("solid", fgColor="5B9BD5")
FILL_OK = PatternFill(fill_type=None)
FILL_YELLOW = PatternFill("solid", fgColor="FFF2CC")
FILL_RED = PatternFill("solid", fgColor="F8CBAD")
FILL_GREEN = PatternFill("solid", fgColor="E2F0D9")
FONT_HEADER = Font(color="FFFFFF", bold=True)
FONT_BOLD = Font(bold=True)
BORDER_THIN = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)

# Rozmiary komentarzy Excela w pikselach. Większe komentarze są istotne,
# bo służą jako audyt kalkulacji wartości statystycznej.
COMMENT_WIDTH_HEADER = 420
COMMENT_HEIGHT_HEADER = 180
COMMENT_WIDTH_NORMAL = 520
COMMENT_HEIGHT_NORMAL = 240
COMMENT_WIDTH_AUDIT = 720
COMMENT_HEIGHT_AUDIT = 380

def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default.copy() if isinstance(default, dict) else default
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(default, dict) and isinstance(data, dict):
            merged = default.copy()
            merged.update(data)
            return merged
        return data
    except Exception:
        return default.copy() if isinstance(default, dict) else default


def save_json(path: Path, data: Any) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

