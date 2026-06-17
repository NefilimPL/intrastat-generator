from __future__ import annotations

from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from .cn import CnResolver
from .config import (
    BORDER_THIN,
    COMMENT_HEIGHT_AUDIT,
    COMMENT_HEIGHT_HEADER,
    COMMENT_HEIGHT_NORMAL,
    COMMENT_WIDTH_AUDIT,
    COMMENT_WIDTH_HEADER,
    COMMENT_WIDTH_NORMAL,
    FILL_GREEN,
    FILL_HEADER,
    FILL_OK,
    FILL_RED,
    FILL_SUBHEADER,
    FILL_YELLOW,
    FONT_BOLD,
    FONT_HEADER,
    OUTPUT_COLUMNS,
    STATUS_MISSING,
    STATUS_OK,
    STATUS_UNCERTAIN,
)
from .models import DictionaryData, IntrastatItem, StatValueResult, TariffEntry
from .naming import make_unique_path
from .text import clean_description, compact_cn, norm_text, safe_int
from .transport import StatisticalValueCalculator
from .version import get_version

def make_comment(text: Any, author: str = "Generator", width: int = COMMENT_WIDTH_NORMAL, height: int = COMMENT_HEIGHT_NORMAL) -> Comment:
    comment = Comment("" if text is None else str(text), author)
    # openpyxl zapisuje rozmiar komentarza do pliku XLSX; Excel respektuje te wartości jako wielkość dymku.
    comment.width = int(width)
    comment.height = int(height)
    return comment


class WorkbookBuilder:
    def __init__(self, dicts: Dict[str, DictionaryData], tariff_entries: List[TariffEntry], resolver: CnResolver, config: Dict[str, Any], route_config: Optional[Dict[str, Any]] = None):
        self.dicts = dicts
        self.tariff_entries = tariff_entries
        self.resolver = resolver
        self.config = config
        self.route_config = route_config or RouteCostManager.default_config()
        self.decisions: List[Dict[str, Any]] = []
        self.warnings: List[str] = []
        self.stat_results: Dict[int, StatValueResult] = {}

    def build(self, items: List[IntrastatItem], declaration_attrs: Dict[str, str], output_path: Path, progress: Optional[Callable[[int, str], None]] = None) -> None:
        if progress:
            progress(45, "Tworzenie skoroszytu XLSX...")
        wb = Workbook()
        ws = wb.active
        ws.title = "ISTAT_IMPORT"

        self._write_main_sheet(ws, items, declaration_attrs, progress)
        self._write_control_sheet(wb)
        self._write_stat_control_sheet(wb, items)
        self._write_settings_sheet(wb, declaration_attrs, output_path)
        self._write_route_cost_sheet(wb)
        self._write_dictionary_sheets(wb)
        self._write_tariff_sheet(wb)
        self._apply_validations(ws, max(len(items) + 1, 2), wb)
        self._finalize_workbook(wb)

        if progress:
            progress(88, "Zapisywanie XLSX...")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path = make_unique_path(output_path)
        wb.save(output_path)

        # Kontrola techniczna: czy openpyxl potrafi ponownie otworzyć wynik.
        try:
            test_wb = load_workbook(output_path, read_only=True, data_only=False)
            test_wb.close()
        except Exception as exc:
            raise RuntimeError(f"XLSX został zapisany, ale nie przeszedł kontroli otwarcia: {exc}") from exc

        if progress:
            progress(100, f"Gotowe: {output_path.name}")

    def _write_main_sheet(self, ws: Any, items: List[IntrastatItem], declaration_attrs: Dict[str, str], progress: Optional[Callable[[int, str], None]]) -> None:
        """Arkusz importowy ma zawierać tylko kolumny wymagane do importu.

        Dane kontrolne wartości statystycznej są przeniesione do komentarzy komórek K
        oraz do osobnego arkusza Kontrola_wartosci_stat.
        """
        ws.append(OUTPUT_COLUMNS)
        for cell in ws[1]:
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER_THIN

        stat_calc = StatisticalValueCalculator(self.config, self.route_config, declaration_attrs)
        self.stat_results = stat_calc.calculate(items)
        self.warnings.extend(stat_calc.warnings)

        for idx, item in enumerate(items, start=2):
            cleaned_desc, warn = clean_description(item.opis)
            for w in warn:
                self.warnings.append(f"Poz {item.poz_id or idx - 1}: {w}")
            decision = self.resolver.resolve(item.opis)
            if not decision.code:
                decision.code = ""
            self.decisions.append({
                "PozId": item.poz_id,
                "Opis": item.opis,
                "Opis po czyszczeniu": cleaned_desc,
                "Kod CN źródłowy XML": item.source_cn,
                "Kod CN finalny": decision.code,
                "Status": decision.status,
                "Pewność %": decision.confidence,
                "Metoda": decision.method,
                "Dopasowanie": decision.matched_text,
                "Uwagi": decision.note,
            })
            stat = self.stat_results.get(idx - 2, StatValueResult(item.statistical_value, "", "", "", "", "", ""))
            ws.append([
                cleaned_desc,
                item.country.upper(),
                item.delivery_terms.upper(),
                item.transaction_type,
                decision.code,
                item.transport_type,
                item.origin_country.upper(),
                item.mass_net,
                item.supplementary_qty,
                item.invoice_value,
                stat.value,
                item.vat_id.upper(),
            ])

            cn_cell = ws.cell(row=idx, column=5)
            if decision.status == STATUS_MISSING:
                cn_cell.fill = FILL_RED
                cn_cell.comment = make_comment(decision.note, width=COMMENT_WIDTH_NORMAL, height=COMMENT_HEIGHT_NORMAL)
            elif decision.status == STATUS_UNCERTAIN:
                cn_cell.fill = FILL_YELLOW
                cn_cell.comment = make_comment(f"Wynik niepewny: {decision.confidence:.1f}%. {decision.note}", width=COMMENT_WIDTH_NORMAL, height=COMMENT_HEIGHT_NORMAL)

            # Komentarz do wartości statystycznej zawiera szczegóły kalkulacji zamiast osobnych kolumn importowych.
            stat_cell = ws.cell(row=idx, column=11)
            stat_comment_lines = []
            if stat.route_name or stat.correction != "" or stat.method or stat.note:
                stat_comment_lines.extend([
                    f"Metoda: {stat.method}",
                    f"Trasa kosztowa: {stat.route_name}",
                    f"Korekta transportu PLN: {stat.correction}",
                    f"Koszt poza PL trasy PLN: {stat.route_total_cost}",
                    f"Udział w koszcie: {stat.share:.4%}" if isinstance(stat.share, (int, float)) else f"Udział w koszcie: {stat.share}",
                ])
                if stat.note:
                    stat_comment_lines.append(f"Uwagi: {stat.note}")
                self._append_comment(stat_cell, "\n".join(stat_comment_lines))

            if stat.status == STATUS_UNCERTAIN:
                stat_cell.fill = FILL_YELLOW
                self._append_comment(stat_cell, stat.note or "Wartość statystyczna wymaga kontroli")

            # Oznaczanie komórek, w których wartości liczbowe są niższe niż 1.
            self._mark_less_than_one(ws.cell(row=idx, column=8), "Masa netto kg")
            self._mark_less_than_one(ws.cell(row=idx, column=10), "Wartość fakturowa PLN")
            self._mark_less_than_one(ws.cell(row=idx, column=11), "Wartość statystyczna PLN")

            if progress and idx % max(1, len(items) // 20 or 1) == 0:
                percent = 45 + int((idx - 1) / max(len(items), 1) * 25)
                progress(min(percent, 70), f"Przetwarzanie pozycji {idx - 1}/{len(items)}...")

        widths = [44, 12, 16, 16, 14, 16, 16, 13, 18, 19, 22, 24]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
        # Nie ustawiamy ws.auto_filter.ref, bo Excel potrafi naprawiać skoroszyt przy jednoczesnym filtrze arkusza i tabeli.

        text_cols = [1, 2, 3, 4, 5, 6, 7, 12]
        int_cols = [8, 9, 10, 11]
        for row in ws.iter_rows(min_row=2, max_row=len(items) + 1, min_col=1, max_col=len(OUTPUT_COLUMNS)):
            for cell in row:
                cell.border = BORDER_THIN
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            for col in text_cols:
                row[col - 1].number_format = "@"
            for col in int_cols:
                row[col - 1].number_format = "0"

        if items:
            table_ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}{len(items)+1}"
            table = Table(displayName="TabelaISTATImport", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)

        ws["C1"].comment = make_comment("Kolumna opcjonalna w Twoim procesie. Ma listę rozwijalną z warunków dostawy, ale może zostać pusta.", width=COMMENT_WIDTH_HEADER, height=COMMENT_HEIGHT_HEADER)
        ws["F1"].comment = make_comment("Kolumna opcjonalna w Twoim procesie. Ma listę rozwijalną z rodzajów transportu, ale może zostać pusta.", width=COMMENT_WIDTH_HEADER, height=COMMENT_HEIGHT_HEADER)
        ws["K1"].comment = make_comment("Wartość statystyczna. Szczegóły kalkulacji są w komentarzach komórek K oraz w arkuszu Kontrola_wartosci_stat.", width=COMMENT_WIDTH_HEADER, height=COMMENT_HEIGHT_HEADER)
        ws["E1"].comment = make_comment("Kod CN bez spacji. Puste czerwone = nie znaleziono pewnego kodu. Żółte = wynik niepewny 80-90% albo reguła wymaga kontroli materiału.", width=COMMENT_WIDTH_HEADER, height=COMMENT_HEIGHT_HEADER)

    def _append_comment(self, cell: Any, text: str) -> None:
        # Zachowujemy podział na linie, bo komentarz służy jako czytelny audyt kalkulacji.
        text = "" if text is None else str(text)
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        text = text.replace(" | ", "\n")
        lines = [norm_text(line) for line in text.split("\n")]
        text = "\n".join(line for line in lines if line)
        if not text:
            return
        if cell.comment and cell.comment.text:
            cell.comment = make_comment(f"{cell.comment.text}\n{text}", width=COMMENT_WIDTH_AUDIT, height=COMMENT_HEIGHT_AUDIT)
        else:
            cell.comment = make_comment(text, width=COMMENT_WIDTH_AUDIT, height=COMMENT_HEIGHT_AUDIT)

    def _mark_less_than_one(self, cell: Any, field_name: str) -> None:
        if cell.value in (None, ""):
            return
        try:
            value = float(str(cell.value).replace(" ", "").replace(",", "."))
        except Exception:
            return
        if value < 1:
            cell.fill = FILL_RED
            self._append_comment(cell, f"UWAGA: {field_name} ma wartość niższą niż 1. Sprawdź, czy pozycja nie jest zerowa albo błędnie przeniesiona.")

    def _write_stat_control_sheet(self, wb: Workbook, items: List[IntrastatItem]) -> None:
        ws = wb.create_sheet("Kontrola_wartosci_stat")
        headers = [
            "PozId", "Opis", "Kod kraju", "VAT kontrahenta", "Masa netto kg", "Wartość fakturowa PLN",
            "Wartość statystyczna PLN", "Korekta transportu PLN", "Koszt poza PL trasy PLN", "Udział w koszcie %",
            "Trasa kosztowa", "Metoda wartości stat.", "Status", "Uwagi"
        ]
        ws.append(headers)
        for idx, item in enumerate(items):
            stat = self.stat_results.get(idx, StatValueResult(item.statistical_value, "", "", "", "", "", ""))
            share_pct = ""
            if isinstance(stat.share, (int, float)):
                share_pct = stat.share
            ws.append([
                item.poz_id,
                item.opis,
                item.country.upper(),
                item.vat_id.upper(),
                item.mass_net,
                item.invoice_value,
                stat.value,
                stat.correction,
                stat.route_total_cost,
                share_pct,
                stat.route_name,
                stat.method,
                stat.status,
                stat.note,
            ])
        self._style_simple_table(ws, len(items) + 1, len(headers), "TabelaKontrolaWartosciStat")
        widths = [10, 50, 12, 24, 14, 20, 22, 22, 24, 16, 24, 70, 14, 100]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
        for row in range(2, len(items) + 2):
            # Kolumny E, F, G: masa/faktura/statystyczna < 1.
            for col in [5, 6, 7]:
                cell = ws.cell(row=row, column=col)
                if cell.value not in (None, ""):
                    try:
                        if float(str(cell.value).replace(" ", "").replace(",", ".")) < 1:
                            cell.fill = FILL_RED
                    except Exception:
                        pass
            if ws.cell(row=row, column=13).value == STATUS_UNCERTAIN:
                for col in range(1, len(headers) + 1):
                    if ws.cell(row=row, column=col).fill == FILL_OK:
                        ws.cell(row=row, column=col).fill = FILL_YELLOW
        for row in range(2, len(items) + 2):
            ws.cell(row=row, column=10).number_format = "0.00%"

    def _write_control_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Kontrola_CN")
        headers = ["PozId", "Opis", "Opis po czyszczeniu", "Kod CN źródłowy XML", "Kod CN finalny", "Status", "Pewność %", "Metoda", "Dopasowanie", "Uwagi"]
        ws.append(headers)
        for d in self.decisions:
            ws.append([d.get(h, "") for h in headers])
        self._style_simple_table(ws, len(self.decisions) + 1, len(headers), "TabelaKontrolaCN")
        widths = [10, 48, 48, 18, 16, 14, 12, 28, 78, 70]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
        for row in range(2, len(self.decisions) + 2):
            status = ws.cell(row=row, column=6).value
            if status == STATUS_MISSING:
                for col in range(1, 11):
                    ws.cell(row=row, column=col).fill = FILL_RED
            elif status == STATUS_UNCERTAIN:
                for col in range(1, 11):
                    ws.cell(row=row, column=col).fill = FILL_YELLOW

    def _write_settings_sheet(self, wb: Workbook, declaration_attrs: Dict[str, str], output_path: Path) -> None:
        ws = wb.create_sheet("Ustawienia")
        rows = [
            ["Nazwa", "Wartość"],
            ["Wygenerowano", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Plik wynikowy", str(output_path)],
            ["Liczba pozycji", len(self.decisions)],
            ["CN próg pewny", self.config.get("cn_confident_threshold")],
            ["CN próg niepewny", self.config.get("cn_uncertain_threshold")],
            ["Domyślne warunki dostawy", self.config.get("default_delivery_terms", "")],
            ["Domyślny rodzaj transakcji", self.config.get("default_transaction_type", "")],
            ["Domyślny rodzaj transportu", self.config.get("default_transport_type", "")],
            ["Tryb wartości statystycznej", self.config.get("statistical_value_mode", "")],
            ["Województwo startowe transportu", self.route_config.get("origin_voivodeship", self.config.get("origin_voivodeship", ""))],
            ["Podział kosztu transportu", self.route_config.get("allocation_basis", self.config.get("transport_allocation_basis", ""))],
            ["Limit kosztu transportu % wartości grupy", self.route_config.get("max_transport_share_pct", self.config.get("transport_max_share_pct", ""))],
            ["Użyj limitu %", self.route_config.get("use_invoice_cap", self.config.get("transport_use_invoice_cap", True))],
            ["Tras kosztowych", len(self.route_config.get("routes", []))],
            ["Słowników wczytanych", len(self.dicts)],
            ["Kodów taryfy wczytanych", len(self.tariff_entries)],
            ["Pozycji z pustym CN", sum(1 for d in self.decisions if d.get("Status") == STATUS_MISSING)],
            ["Pozycji z niepewnym CN", sum(1 for d in self.decisions if d.get("Status") == STATUS_UNCERTAIN)],
        ]
        for k, v in declaration_attrs.items():
            rows.append([f"Deklaracja.{k}", v])
        if self.warnings:
            rows.append(["", ""])
            rows.append(["OSTRZEŻENIA", ""])
            for w in self.warnings:
                rows.append(["", w])
        for r in rows:
            ws.append(r)
        self._style_simple_table(ws, len(rows), 2, "TabelaUstawienia")
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 96

    def _write_route_cost_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Koszty_transportu")
        ws.append(["Kraj", "Część kraju/strefa", "Koszt poza PL 1 TIR PLN", "Liczba transportów", "Max % faktur", "Domyślna", "Aktywna", "Uwagi"])
        for r in self.route_config.get("routes", []):
            ws.append([
                norm_text(r.get("country", "")).upper(),
                norm_text(r.get("zone", "STANDARD")).upper(),
                safe_float(r.get("foreign_cost_pln", 0.0)),
                safe_float(r.get("truck_count", 1.0)),
                safe_float(r.get("max_correction_pct", self.route_config.get("max_transport_share_pct", self.config.get("transport_max_share_pct", 8.0))), 8.0),
                yes_no(r.get("default", False)),
                yes_no(r.get("active", True)),
                norm_text(r.get("note", "")),
            ])
        self._style_simple_table(ws, max(2, len(self.route_config.get("routes", [])) + 1), 8, "TabelaKosztyTransportu")
        widths = [10, 20, 24, 18, 14, 12, 12, 70]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
        ws["A1"].comment = make_comment("Tabela kontrolna użyta do wyliczenia wartości statystycznej. Edytuj ją w GUI, nie w arkuszu wynikowym.", width=COMMENT_WIDTH_HEADER, height=COMMENT_HEIGHT_HEADER)
        ws["E1"].comment = make_comment("Maksymalny koszt transportu użyty w kalkulacji jako procent sumy faktur dla danej grupy kraju/strefy.", width=COMMENT_WIDTH_HEADER, height=COMMENT_HEIGHT_HEADER)


    def _write_dictionary_sheets(self, wb: Workbook) -> None:
        for code in sorted(self.dicts.keys()):
            d = self.dicts[code]
            ws = wb.create_sheet(f"SLW_{code}")
            ws.append(["Kod", "Opis", "OpisEN", "WaznyOd", "WaznyDo", "Nazwa słownika", "Plik"])
            for row in d.rows:
                ws.append([
                    row.get("Kod", ""), row.get("Opis", ""), row.get("OpisEN", ""), row.get("WaznyOd", ""), row.get("WaznyDo", ""), d.name, d.path.name
                ])
            self._style_simple_table(ws, len(d.rows) + 1, 7, f"TabelaSLW{code}")
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 72
            ws.column_dimensions["C"].width = 56
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 14
            ws.column_dimensions["F"].width = 44
            ws.column_dimensions["G"].width = 32
            ws.freeze_panes = "A2"
            if bool(self.config.get("hide_dictionary_sheets", False)):
                ws.sheet_state = "hidden"

    def _write_tariff_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Taryfa_CN")
        ws.append(["Kod CN", "Kod z odstępami", "Opis", "Ścieżka taryfy"])
        rows = [e for e in self.tariff_entries if e.code.startswith("94") or e.code.startswith("99")]
        for e in rows:
            ws.append([e.code, e.spaced_code, e.description, e.path_text])
        self._style_simple_table(ws, len(rows) + 1, 4, "TabelaTaryfaCN")
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 58
        ws.column_dimensions["D"].width = 100
        ws.freeze_panes = "A2"

    def _style_simple_table(self, ws: Any, rows: int, cols: int, table_name: str) -> None:
        for cell in ws[1]:
            cell.fill = FILL_SUBHEADER
            cell.font = FONT_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER_THIN
        for row in ws.iter_rows(min_row=2, max_row=max(rows, 2), min_col=1, max_col=cols):
            for cell in row:
                cell.border = BORDER_THIN
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.column in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
                    cell.number_format = "@"
        if rows >= 2:
            try:
                ref = f"A1:{get_column_letter(cols)}{rows}"
                table = Table(displayName=table_name, ref=ref)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                ws.add_table(table)
            except Exception:
                pass

    def _apply_validations(self, ws: Any, max_row: int, wb: Workbook) -> None:
        # Formuły listy zapisujemy bez znaku '='. Taki zapis jest bezpieczniejszy dla Excela.
        self._add_list_validation(ws, f"B2:B{max_row}", self._range_for_dict("049") or self._range_for_dict("007"), False,
                                  "Kod kraju UE ze słownika XML 049/007.")
        self._add_list_validation(ws, f"C2:C{max_row}", self._range_for_dict("002"), True,
                                  "Warunki dostawy. Pole może być puste.")
        self._add_list_validation(ws, f"D2:D{max_row}", self._range_for_dict("004"), False,
                                  "Rodzaj transakcji ze słownika XML 004.")
        tariff_last = max(2, len([e for e in self.tariff_entries if e.code.startswith("94") or e.code.startswith("99")]) + 1)
        if "Taryfa_CN" in wb.sheetnames and tariff_last >= 2:
            self._add_list_validation(ws, f"E2:E{max_row}", f"{quote_sheetname('Taryfa_CN')}!$A$2:$A${tariff_last}", True,
                                      "Kod CN z arkusza Taryfa_CN. Czerwone puste pola wymagają ręcznego wyboru.")
        self._add_list_validation(ws, f"F2:F{max_row}", self._range_for_dict("005"), True,
                                  "Rodzaj transportu. Pole może być puste.")
        self._add_list_validation(ws, f"G2:G{max_row}", self._range_for_dict("007"), True,
                                  "Kod kraju pochodzenia ze słownika XML 007.")

    def _range_for_dict(self, code: str) -> str:
        d = self.dicts.get(code)
        if not d or not d.rows:
            return ""
        sheet_name = f"SLW_{code}"
        last = len(d.rows) + 1
        return f"{quote_sheetname(sheet_name)}!$A$2:$A${last}"

    def _add_list_validation(self, ws: Any, cell_range: str, formula_range: str, allow_blank: bool, prompt: str) -> None:
        if not formula_range:
            return
        dv = DataValidation(type="list", formula1=formula_range, allow_blank=allow_blank)
        dv.error = "Wartość spoza słownika. Wybierz kod z listy albo popraw słownik."
        dv.errorTitle = "Błędna wartość"
        dv.prompt = prompt
        dv.promptTitle = "Słownik"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        # showDropDown=False oznacza w OOXML, że strzałka listy ma być widoczna w Excelu.
        dv.showDropDown = False
        ws.add_data_validation(dv)
        dv.add(cell_range)

    def _finalize_workbook(self, wb: Workbook) -> None:
        wb.properties.creator = f"Generator INTRASTAT XLSX {get_version()}"
        wb.properties.title = "INTRASTAT import XLSX"
        wb.properties.subject = "Wygenerowany plik importu ist@t2"
        wb.active = 0


