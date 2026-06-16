#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generator XLSX INTRASTAT z GUI.

Założenia wersji v2:
- Brak cn_overrides.json. Kod CN jest wyznaczany automatycznie z reguł + taryfy + fuzzy.
- Jeżeli nie da się dobrać kodu z dostępnych pozycji taryfy, komórka CN zostaje pusta i czerwona.
- Jeżeli wynik jest niepewny, domyślnie 80-89.99%, komórka CN jest żółta.
- Kod CN w XLSX jest zapisywany bez spacji, np. 94036010.
- Słowniki XML są wczytywane dynamicznie i używane do walidacji danych w Excelu.
- GUI ma listy rozwijalne dla domyślnych wartości, jeżeli odpowiednie słowniki istnieją.
- Wynik XLSX jest zapisywany do folderu wygenerowane_xlsx obok programu.

Instalacja:
    pip install -r requirements_intrastat_generator_v2.txt

Uruchomienie:
    python intrastat_generator_gui_v2.py

Tryb bez GUI:
    python intrastat_generator_gui_v2.py --input "intrastat.xml" --tariff "taryfa.txt" --no-gui
"""

from __future__ import annotations

import argparse
import json
import os
import queue
import re
import sys
import threading
import time
import traceback
import unicodedata
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except Exception:
    tk = None  # type: ignore
    filedialog = None  # type: ignore
    messagebox = None  # type: ignore
    ttk = None  # type: ignore

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD  # type: ignore
    HAS_DND = True
except Exception:
    DND_FILES = None  # type: ignore
    TkinterDnD = None  # type: ignore
    HAS_DND = False

try:
    from rapidfuzz import fuzz, process  # type: ignore
    HAS_RAPIDFUZZ = True
except Exception:
    HAS_RAPIDFUZZ = False
    fuzz = None  # type: ignore
    process = None  # type: ignore
    import difflib

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter, quote_sheetname
except Exception as exc:
    print("Brak biblioteki openpyxl. Zainstaluj: pip install openpyxl", file=sys.stderr)
    raise exc

APP_NAME = "Generator INTRASTAT XLSX v2"
CONFIG_FILE = "config.json"
OUTPUT_DIR_NAME = "wygenerowane_xlsx"
DICT_DIR_NAME = "slowniki"
LOG_DIR_NAME = "logi"

FORBIDDEN_DESC_CHARS = ["&", '"', "'", "<", ">", ";"]

OUTPUT_COLUMNS = [
    "Opis towaru",                # A - tekst
    "Kod kraju",                 # B - tekst
    "Warunki dostawy",           # C - tekst, opcjonalne
    "Rodzaj transakcji",         # D - liczba wg instrukcji, ale zapis tekstowy pomaga nie gubić zer
    "Kod towaru CN",             # E - liczba wg instrukcji, zapis tekstowy chroni format 8 cyfr
    "Rodzaj transportu",         # F - liczba wg instrukcji, opcjonalne
    "Kraj pochodzenia",          # G - tekst
    "Masa netto kg",             # H - liczba całkowita
    "Ilość w jedn. uzup.",       # I - liczba całkowita/opcjonalna
    "Wartość fakturowa PLN",     # J - liczba całkowita
    "Wartość statystyczna PLN",  # K - liczba całkowita/opcjonalna
    "VAT kontrahenta",           # L - tekst
]

DEFAULT_CONFIG: Dict[str, Any] = {
    "tariff_path": "",
    "dict_dir": DICT_DIR_NAME,
    "output_dir": OUTPUT_DIR_NAME,
    "default_delivery_terms": "",
    "default_transaction_type": "11",
    "default_transport_type": "",
    "statistical_value_mode": "blank",  # blank | copy_invoice_when_required | copy_invoice_always
    "auto_open_output_folder": False,
    "cn_confident_threshold": 90.0,
    "cn_uncertain_threshold": 80.0,
    "hide_dictionary_sheets": False,
}

STATUS_OK = "OK"
STATUS_UNCERTAIN = "NIEPEWNY"
STATUS_MISSING = "BRAK"

FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
FILL_SUBHEADER = PatternFill("solid", fgColor="5B9BD5")
FILL_OK = PatternFill(fill_type=None)
FILL_YELLOW = PatternFill("solid", fgColor="FFF2CC")
FILL_RED = PatternFill("solid", fgColor="F8CBAD")
FILL_GREEN = PatternFill("solid", fgColor="E2F0D9")
FONT_HEADER = Font(color="FFFFFF", bold=True)
FONT_BOLD = Font(bold=True)
BORDER_THIN = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def get_app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def resolve_path(value: str | Path, base_dir: Path) -> Path:
    p = Path(value)
    if p.is_absolute():
        return p
    return base_dir / p


def ensure_dirs(base_dir: Path, config: Dict[str, Any]) -> Dict[str, Path]:
    paths = {
        "base": base_dir,
        "dict": resolve_path(config.get("dict_dir") or DICT_DIR_NAME, base_dir),
        "output": resolve_path(config.get("output_dir") or OUTPUT_DIR_NAME, base_dir),
        "logs": base_dir / LOG_DIR_NAME,
    }
    for p in paths.values():
        p.mkdir(parents=True, exist_ok=True)
    return paths


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S_%f")


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default.copy() if isinstance(default, dict) else default
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(default, dict) and isinstance(data, dict):
            merged = default.copy()
            merged.update(data)
            return merged
        return data
    except Exception:
        return default.copy() if isinstance(default, dict) else default


def save_json(path: Path, data: Any) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def log_exception(base_dir: Path, exc: BaseException) -> Path:
    log_dir = base_dir / LOG_DIR_NAME
    log_dir.mkdir(parents=True, exist_ok=True)
    p = log_dir / f"blad_{now_stamp()}.log"
    with p.open("w", encoding="utf-8") as f:
        f.write("".join(traceback.format_exception(type(exc), exc, exc.__traceback__)))
    return p


def strip_ns(tag: str) -> str:
    return tag.split("}", 1)[-1] if "}" in tag else tag


def norm_text(value: Any) -> str:
    s = "" if value is None else str(value)
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def norm_key(value: Any) -> str:
    s = norm_text(value).upper()
    s = s.translate(str.maketrans({"Ł": "L", "ł": "L", "Đ": "D", "đ": "D", "Ø": "O", "ø": "O"}))
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"[^A-Z0-9./()+\- ]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def tokens(value: str) -> set[str]:
    s = norm_key(value)
    return set(re.findall(r"[A-Z0-9]+", s))


def compact_cn(value: Any) -> str:
    digits = re.sub(r"\D", "", "" if value is None else str(value))
    return digits[:8] if len(digits) >= 8 else digits


def clean_description(desc: str) -> Tuple[str, List[str]]:
    warnings: List[str] = []
    s = norm_text(desc)
    for ch in FORBIDDEN_DESC_CHARS:
        if ch in s:
            s = s.replace(ch, " ")
            warnings.append(f"Usunięto niedozwolony znak {ch!r} z opisu")
    s = re.sub(r"\s+", " ", s).strip()
    if len(s) > 255:
        s = s[:255]
        warnings.append("Opis ucięty do 255 znaków")
    return s, warnings


def safe_int(value: Any, blank_if_missing: bool = True) -> Any:
    s = norm_text(value)
    if s == "":
        return "" if blank_if_missing else 0
    s = s.replace(" ", "").replace(",", ".")
    try:
        return int(round(float(s)))
    except Exception:
        return "" if blank_if_missing else 0


def path_from_drop_data(root: tk.Tk, data: str) -> str:  # type: ignore[name-defined]
    try:
        items = root.tk.splitlist(data)
        if items:
            return str(items[0])
    except Exception:
        pass
    return data.strip().strip("{}").strip('"')


def make_unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for i in range(1, 1000):
        candidate = path.with_name(f"{stem}_{i}{suffix}")
        if not candidate.exists():
            return candidate
    return path.with_name(f"{stem}_{now_stamp()}{suffix}")


@dataclass
class DictionaryData:
    code: str
    name: str
    date: str
    path: Path
    rows: List[Dict[str, str]]

    def codes(self) -> List[str]:
        return [r.get("Kod", "") for r in self.rows if r.get("Kod")]


@dataclass
class TariffEntry:
    code: str
    spaced_code: str
    description: str
    path_text: str


@dataclass
class IntrastatItem:
    poz_id: str
    opis: str
    country: str
    delivery_terms: str
    transaction_type: str
    source_cn: str
    transport_type: str
    origin_country: str
    mass_net: Any
    supplementary_qty: Any
    invoice_value: Any
    statistical_value: Any
    vat_id: str
    attrs: Dict[str, str]


@dataclass
class CnDecision:
    code: str
    status: str
    confidence: float
    method: str
    matched_text: str
    note: str


class DictionaryLoader:
    def __init__(self, base_dir: Path, dict_dir: Path):
        self.base_dir = base_dir
        self.dict_dir = dict_dir

    def discover_paths(self) -> List[Path]:
        found: Dict[str, Path] = {}
        for folder in [self.dict_dir, self.base_dir]:
            if not folder.exists():
                continue
            for p in folder.glob("*.xml"):
                if p.name.lower().startswith("slownik"):
                    found[str(p.resolve()).lower()] = p
        return sorted(found.values(), key=lambda p: p.name.lower())

    def load(self) -> Dict[str, DictionaryData]:
        latest: Dict[str, DictionaryData] = {}
        for path in self.discover_paths():
            data = self._parse_one(path)
            if not data:
                continue
            old = latest.get(data.code)
            if old is None or (data.date or "0000-00-00") >= (old.date or "0000-00-00"):
                latest[data.code] = data
        return latest

    def _parse_one(self, path: Path) -> Optional[DictionaryData]:
        try:
            root = ET.parse(path).getroot()
        except Exception:
            return None
        if strip_ns(root.tag) != "Slownik":
            return None
        code = norm_text(root.attrib.get("Kod", ""))
        name = norm_text(root.attrib.get("Nazwa", "") or root.attrib.get("NazwaEN", ""))
        date = norm_text(root.attrib.get("DataEdycji", ""))
        rows: List[Dict[str, str]] = []
        for pos in root.iter():
            if strip_ns(pos.tag) != "Pozycja":
                continue
            rows.append({
                "Kod": norm_text(pos.attrib.get("Kod", "")),
                "Opis": norm_text(pos.attrib.get("Opis", "")),
                "OpisEN": norm_text(pos.attrib.get("OpisEN", "")),
                "WaznyOd": norm_text(pos.attrib.get("WaznyOd", "")),
                "WaznyDo": norm_text(pos.attrib.get("WaznyDo", "")),
            })
        return DictionaryData(code=code, name=name, date=date, path=path, rows=rows)


class TariffLoader:
    CODE_RE = re.compile(r"^\s*((?:\d{4})(?:\s+\d{2}){1,2})\s*-\s*(.+?)\s*$")

    def __init__(self, tariff_path: Path):
        self.tariff_path = tariff_path

    def load(self) -> List[TariffEntry]:
        entries: List[TariffEntry] = []
        context: List[Tuple[int, str]] = []
        with self.tariff_path.open("r", encoding="utf-8", errors="replace") as f:
            for raw in f:
                line = raw.rstrip("\n\r")
                if not line.strip():
                    continue
                indent = len(line) - len(line.lstrip("\t "))
                stripped = line.strip()
                match = self.CODE_RE.match(line)
                if match:
                    spaced = re.sub(r"\s+", " ", match.group(1).strip())
                    code = compact_cn(spaced)
                    desc = re.sub(r"\s*\[[^\]]*\]\s*$", "", match.group(2).strip())
                    if len(code) == 8:
                        path_text = " > ".join([c[1] for c in context[-7:]] + [desc])
                        entries.append(TariffEntry(code=code, spaced_code=spaced, description=desc, path_text=path_text))
                    self._push_context(context, indent, f"{spaced} {desc}")
                else:
                    if len(stripped) <= 150 and not stripped.startswith("TARYFA"):
                        self._push_context(context, indent, stripped)
        # Deduplikacja po kodzie i ścieżce, bo w niektórych plikach taryfa powtarza sekcje.
        unique: Dict[str, TariffEntry] = {}
        for e in entries:
            if e.code not in unique or len(e.path_text) > len(unique[e.code].path_text):
                unique[e.code] = e
        return sorted(unique.values(), key=lambda x: x.code)

    @staticmethod
    def _push_context(context: List[Tuple[int, str]], indent: int, text: str) -> None:
        while context and context[-1][0] >= indent:
            context.pop()
        context.append((indent, text))


class CnResolver:
    """Dobór CN bez ręcznego słownika produktów.

    Kolejność:
    1. Reguły klasyfikacyjne po typie towaru i materiale opisanym/założonym w nazwie.
    2. Fuzzy matching po pozycjach taryfy działu 94 oraz dodatkowych kodach 99.
    3. Brak kodu, jeżeli wynik nie spełnia progu lub kodu nie ma w taryfie.
    """

    def __init__(self, tariff_entries: List[TariffEntry], confident_threshold: float, uncertain_threshold: float):
        self.tariff_entries = tariff_entries
        self.confident_threshold = float(confident_threshold)
        self.uncertain_threshold = float(uncertain_threshold)
        self.valid_codes = {e.code for e in tariff_entries}
        self.relevant_entries = [e for e in tariff_entries if e.code.startswith("94") or e.code.startswith("99")]
        self._fuzzy_texts = [self._normalize_for_fuzzy(f"{e.path_text} {e.description}") for e in self.relevant_entries]

    @staticmethod
    def _normalize_for_fuzzy(value: str) -> str:
        s = norm_key(value)
        s = re.sub(r"[^A-Z0-9 ]+", " ", s)
        return re.sub(r"\s+", " ", s).strip()

    def resolve(self, description: str) -> CnDecision:
        rule = self._rule_based(description)
        if rule:
            code, confidence, method, note = rule
            if code in self.valid_codes:
                return self._with_status(code, confidence, method, self._entry_text(code), note)
            return CnDecision("", STATUS_MISSING, 0.0, method, "", f"Reguła wskazała {code}, ale kodu nie ma w wczytanej taryfie")

        fuzzy_decision = self._tariff_fuzzy(description)
        if fuzzy_decision.code and fuzzy_decision.confidence >= self.uncertain_threshold and fuzzy_decision.code in self.valid_codes:
            return self._with_status(fuzzy_decision.code, fuzzy_decision.confidence, fuzzy_decision.method, fuzzy_decision.matched_text, fuzzy_decision.note)

        best = fuzzy_decision
        if best.code:
            return CnDecision("", STATUS_MISSING, best.confidence, "brak pewnego dopasowania", best.matched_text,
                              f"Najlepsze dopasowanie {best.code} miało tylko {best.confidence:.1f}%, poniżej progu {self.uncertain_threshold:.1f}%")
        return CnDecision("", STATUS_MISSING, 0.0, "brak", "", "Nie znaleziono żadnego pasującego kodu w taryfie")

    def _with_status(self, code: str, confidence: float, method: str, matched_text: str, note: str) -> CnDecision:
        status = STATUS_OK if confidence >= self.confident_threshold else STATUS_UNCERTAIN
        return CnDecision(code, status, round(float(confidence), 2), method, matched_text, note)

    def _entry_text(self, code: str) -> str:
        for e in self.tariff_entries:
            if e.code == code:
                return e.path_text
        return ""

    def _rule_based(self, description: str) -> Optional[Tuple[str, float, str, str]]:
        s = norm_key(description)
        t = tokens(description)

        # Siedzenia. Tu bez kartoteki materiałowej część decyzji pozostaje niepewna.
        if "FOTEL" in t and "OBROTOWY" in t:
            return "94013900", 86.0, "reguła: fotel obrotowy", "Siedzenie obrotowe, materiał ramy niepewny; wymaga kontroli, dlatego żółty wynik"
        if "FOTEL" in t and "OGRODOWY" in t:
            return "94017100", 86.0, "reguła: fotel ogrodowy", "Założono siedzenie tapicerowane/z poduszką na ramie metalowej; wymaga kontroli"
        if "HOKER" in t or "KRZESLO" in t:
            if any(x in s for x in ["WELUR", "WELUROWE", "TAPIC", "TK.", "TK ", "ECO", "BEZ", "SZARE", "CZARNE"]):
                return "94017100", 86.0, "reguła: krzesło/hoker", "Założono siedzenie tapicerowane na ramie metalowej; wymaga kontroli materiału ramy"
            return "94017900", 82.0, "reguła: krzesło/hoker", "Brak informacji o tapicerce; założono pozostałe siedzenie na ramie metalowej"
        if "TABORET" in t or "LAWKA" in t:
            if any(x in s for x in ["TK", "ECO", "SAFARI", "WELUR", "TAPIC"]):
                return "94016100", 86.0, "reguła: taboret/ławka", "Założono siedzenie tapicerowane na ramie drewnianej; wymaga kontroli"
            return "94016900", 82.0, "reguła: taboret/ławka", "Brak informacji o tapicerce; założono pozostałe siedzenie drewniane"
        if "FOTEL" in t:
            return "", 0.0, "reguła: fotel bez danych", "Nie można bezpiecznie ustalić typu fotela/ramy z samego opisu"

        # Meble kuchenne.
        if "KUCH" in s or "LIVIA" in t:
            return "94034090", 92.0, "reguła: mebel kuchenny", "Element/mebel kuchenny drewniany lub z płyty"

        # Meble biurowe.
        if "BIURKO" in t:
            return "94033011", 95.0, "reguła: biurko", "Biurko drewniane lub z płyty"

        # Meble sypialniane.
        if "SZAFA" in t or ("SZAFKA" in t and "NOCNA" in t) or ("STOLIK" in t and "NOCNY" in t):
            return "94035000", 94.0, "reguła: sypialnia", "Mebel drewniany/z płyty w rodzaju stosowanych w sypialni"

        # Przedpokój, wieszak, szafka na buty, szafka z lustrem i ogólne cabinet/high cabinet.
        if "PRZEDPOKOJ" in t or "WIESZAK" in t or "BUTY" in t or "HALL" in t or "LUSTREM" in t or "CABINET" in t:
            return "94036090", 92.0, "reguła: pozostały mebel drewniany", "Pozostały mebel drewniany/z płyty"

        # Szafka bez oznaczenia kuchni/sypialni traktowana jako mebel pokojowy/salonowy.
        if "SZAFKA" in t:
            return "94036010", 92.0, "reguła: szafka pokojowa", "Szafka drewniana/z płyty, bez cech kuchni/sypialni/przedpokoju"

        # Salon, pokój dzienny, jadalnia.
        living_words = {"RTV", "KOMODA", "LAWA", "STOL", "STOLIK", "MEBLOSCIANKA", "WITRYNA", "REGAL"}
        if living_words.intersection(t):
            return "94036010", 95.0, "reguła: salon/jadalnia", "Mebel drewniany/z płyty w rodzaju stosowanych w pokojach stołowych i salonach"

        # Typowy zapis mebli skrzyniowych bez słowa KOMODA, np. 2D1S, 3D3S.
        if re.search(r"\b\d+D(?:\d*S)?\b|\b\d+S\b", s):
            return "94036010", 90.0, "reguła: układ drzwi/szuflad", "Opis zawiera układ typu 2D1S/3D3S, czyli typowy mebel skrzyniowy pokojowy"

        # Angielskie nazwy szafek, które nie mówią o pomieszczeniu.
        if "CABINET" in t:
            return "94036090", 88.0, "reguła: cabinet", "Pozostały mebel drewniany/z płyty, opis nie wskazuje jednoznacznie pomieszczenia"

        return None

    def _tariff_fuzzy(self, description: str) -> CnDecision:
        if not self.relevant_entries:
            return CnDecision("", STATUS_MISSING, 0.0, "brak taryfy", "", "Nie załadowano pozycji działu 94/99 z taryfy")
        query = self._normalize_for_fuzzy(description)
        if HAS_RAPIDFUZZ:
            result = process.extractOne(query, self._fuzzy_texts, scorer=fuzz.WRatio)  # type: ignore[union-attr]
            if result:
                _text, score, idx = result
                entry = self.relevant_entries[idx]
                return CnDecision(entry.code, STATUS_UNCERTAIN, float(score), "fuzzy taryfa", entry.path_text, "Najlepsze podobieństwo do opisu taryfy")
        else:
            best_idx = 0
            best_score = -1.0
            for i, txt in enumerate(self._fuzzy_texts):
                score = difflib.SequenceMatcher(None, query, txt).ratio() * 100
                if score > best_score:
                    best_score = score
                    best_idx = i
            entry = self.relevant_entries[best_idx]
            return CnDecision(entry.code, STATUS_UNCERTAIN, best_score, "fuzzy taryfa", entry.path_text, "Najlepsze podobieństwo do opisu taryfy")
        return CnDecision("", STATUS_MISSING, 0.0, "brak", "", "Nie znaleziono dopasowania")


class IntrastatXmlParser:
    def __init__(self, path: Path, config: Dict[str, Any], dicts: Dict[str, DictionaryData]):
        self.path = path
        self.config = config
        self.dicts = dicts
        self.declaration_attrs: Dict[str, str] = {}

    @staticmethod
    def _get_attr(attrs: Dict[str, str], candidates: Sequence[str], default: str = "") -> str:
        for c in candidates:
            if c in attrs and norm_text(attrs[c]) != "":
                return norm_text(attrs[c])
        low = {k.lower(): v for k, v in attrs.items()}
        for c in candidates:
            v = low.get(c.lower())
            if v is not None and norm_text(v) != "":
                return norm_text(v)
        return default

    def parse(self) -> List[IntrastatItem]:
        root = ET.parse(self.path).getroot()
        declaration = None
        for elem in root.iter():
            if strip_ns(elem.tag) == "Deklaracja":
                declaration = elem
                break
        if declaration is not None:
            self.declaration_attrs = {k: norm_text(v) for k, v in declaration.attrib.items()}

        items: List[IntrastatItem] = []
        for elem in root.iter():
            if strip_ns(elem.tag) != "Towar":
                continue
            attrs = {k: norm_text(v) for k, v in elem.attrib.items()}
            source_cn = compact_cn(self._get_attr(attrs, ["KodTowarowy", "KodCN", "CN", "KodTowaru"]))
            transaction = self._get_attr(attrs, ["RodzajTransakcji", "KodRodzajuTransakcji"], self.config.get("default_transaction_type", ""))
            invoice_value = safe_int(self._get_attr(attrs, ["WartoscFaktury", "WartoscFakturowa", "WartoscFakturyPLN"]))
            stat_value_xml = safe_int(self._get_attr(attrs, ["WartoscStatystyczna", "WartoscStatystycznaPLN"]))
            stat_value = self._statistical_value(stat_value_xml, invoice_value, transaction, source_cn)

            items.append(IntrastatItem(
                poz_id=self._get_attr(attrs, ["PozId", "Lp", "Pozycja"]),
                opis=self._get_attr(attrs, ["OpisTowaru", "Opis", "NazwaTowaru"]),
                country=self._get_attr(attrs, ["KrajPrzeznaczeniaWysylki", "KrajPrzeznaczenia", "KrajWysylki", "KrajWywozu"]),
                delivery_terms=self._get_attr(attrs, ["WarunkiDostawy", "KodWarunkowDostawy"], self.config.get("default_delivery_terms", "")),
                transaction_type=transaction,
                source_cn=source_cn,
                transport_type=self._get_attr(attrs, ["RodzajTransportu", "KodRodzajuTransportu"], self.config.get("default_transport_type", "")),
                origin_country=self._get_attr(attrs, ["KrajPochodzenia", "KodKrajuPochodzenia"]),
                mass_net=safe_int(self._get_attr(attrs, ["MasaNetto", "MasaNettoKg"]), blank_if_missing=False),
                supplementary_qty=safe_int(self._get_attr(attrs, ["IloscJednostekUzupelniajacych", "IloscJmUzupelniajaca", "IloscUzupelniajaca"])),
                invoice_value=invoice_value,
                statistical_value=stat_value,
                vat_id=self._get_attr(attrs, ["IdKontrahenta", "VatKontrahenta", "NumerVatKontrahenta", "NIPKontrahenta"]),
                attrs=attrs,
            ))
        return items

    def _statistical_value(self, xml_value: Any, invoice_value: Any, transaction_type: str, cn: str) -> Any:
        if xml_value != "":
            return xml_value
        mode = self.config.get("statistical_value_mode", "blank")
        if mode == "copy_invoice_always":
            return invoice_value
        if mode == "copy_invoice_when_required":
            required_by_transaction = transaction_type in set(self.dicts.get("191", DictionaryData("", "", "", Path(), [])).codes())
            required_by_cn = cn in set(self.dicts.get("190", DictionaryData("", "", "", Path(), [])).codes())
            if required_by_transaction or required_by_cn:
                return invoice_value
        return ""


class WorkbookBuilder:
    def __init__(self, dicts: Dict[str, DictionaryData], tariff_entries: List[TariffEntry], resolver: CnResolver, config: Dict[str, Any]):
        self.dicts = dicts
        self.tariff_entries = tariff_entries
        self.resolver = resolver
        self.config = config
        self.decisions: List[Dict[str, Any]] = []
        self.warnings: List[str] = []

    def build(self, items: List[IntrastatItem], declaration_attrs: Dict[str, str], output_path: Path, progress: Optional[Callable[[int, str], None]] = None) -> None:
        if progress:
            progress(45, "Tworzenie skoroszytu XLSX...")
        wb = Workbook()
        ws = wb.active
        ws.title = "ISTAT_IMPORT"

        self._write_main_sheet(ws, items, progress)
        self._write_control_sheet(wb)
        self._write_settings_sheet(wb, declaration_attrs, output_path)
        self._write_dictionary_sheets(wb)
        self._write_tariff_sheet(wb)
        self._apply_validations(ws, max(len(items) + 1, 2), wb)
        self._finalize_workbook(wb)

        if progress:
            progress(88, "Zapisywanie XLSX...")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path = make_unique_path(output_path)
        wb.save(output_path)

        # Kontrola techniczna: czy openpyxl potrafi ponownie otworzyć wynik.
        try:
            test_wb = load_workbook(output_path, read_only=True, data_only=False)
            test_wb.close()
        except Exception as exc:
            raise RuntimeError(f"XLSX został zapisany, ale nie przeszedł kontroli otwarcia: {exc}") from exc

        if progress:
            progress(100, f"Gotowe: {output_path.name}")

    def _write_main_sheet(self, ws: Any, items: List[IntrastatItem], progress: Optional[Callable[[int, str], None]]) -> None:
        ws.append(OUTPUT_COLUMNS)
        for cell in ws[1]:
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER_THIN

        for idx, item in enumerate(items, start=2):
            cleaned_desc, warn = clean_description(item.opis)
            for w in warn:
                self.warnings.append(f"Poz {item.poz_id or idx - 1}: {w}")
            decision = self.resolver.resolve(item.opis)
            if not decision.code:
                decision.code = ""
            self.decisions.append({
                "PozId": item.poz_id,
                "Opis": item.opis,
                "Opis po czyszczeniu": cleaned_desc,
                "Kod CN źródłowy XML": item.source_cn,
                "Kod CN finalny": decision.code,
                "Status": decision.status,
                "Pewność %": decision.confidence,
                "Metoda": decision.method,
                "Dopasowanie": decision.matched_text,
                "Uwagi": decision.note,
            })
            ws.append([
                cleaned_desc,
                item.country.upper(),
                item.delivery_terms.upper(),
                item.transaction_type,
                decision.code,
                item.transport_type,
                item.origin_country.upper(),
                item.mass_net,
                item.supplementary_qty,
                item.invoice_value,
                item.statistical_value,
                item.vat_id.upper(),
            ])
            cn_cell = ws.cell(row=idx, column=5)
            if decision.status == STATUS_MISSING:
                cn_cell.fill = FILL_RED
                cn_cell.comment = Comment(decision.note, "Generator")
            elif decision.status == STATUS_UNCERTAIN:
                cn_cell.fill = FILL_YELLOW
                cn_cell.comment = Comment(f"Wynik niepewny: {decision.confidence:.1f}%. {decision.note}", "Generator")

            if progress and idx % max(1, len(items) // 20 or 1) == 0:
                percent = 45 + int((idx - 1) / max(len(items), 1) * 25)
                progress(min(percent, 70), f"Przetwarzanie pozycji {idx - 1}/{len(items)}...")

        widths = [44, 12, 16, 16, 14, 16, 16, 13, 18, 19, 22, 24]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
        # Nie ustawiamy ws.auto_filter.ref, bo Excel potrafi naprawiać skoroszyt przy jednoczesnym filtrze arkusza i tabeli.

        text_cols = [1, 2, 3, 4, 5, 6, 7, 12]
        int_cols = [8, 9, 10, 11]
        for row in ws.iter_rows(min_row=2, max_row=len(items) + 1, min_col=1, max_col=12):
            for cell in row:
                cell.border = BORDER_THIN
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            for col in text_cols:
                row[col - 1].number_format = "@"
            for col in int_cols:
                row[col - 1].number_format = "0"

        if items:
            table = Table(displayName="TabelaISTATImport", ref=f"A1:L{len(items)+1}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)

        ws["C1"].comment = Comment("Kolumna opcjonalna w Twoim procesie. Ma listę rozwijalną z warunków dostawy, ale może zostać pusta.", "Generator")
        ws["F1"].comment = Comment("Kolumna opcjonalna w Twoim procesie. Ma listę rozwijalną z rodzajów transportu, ale może zostać pusta.", "Generator")
        ws["K1"].comment = Comment("Kolumna opcjonalna. Domyślnie pusta, chyba że XML lub ustawienia programu podają wartość.", "Generator")
        ws["E1"].comment = Comment("Kod CN bez spacji. Puste czerwone = nie znaleziono pewnego kodu. Żółte = wynik niepewny 80-90% albo reguła wymaga kontroli materiału.", "Generator")

    def _write_control_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Kontrola_CN")
        headers = ["PozId", "Opis", "Opis po czyszczeniu", "Kod CN źródłowy XML", "Kod CN finalny", "Status", "Pewność %", "Metoda", "Dopasowanie", "Uwagi"]
        ws.append(headers)
        for d in self.decisions:
            ws.append([d.get(h, "") for h in headers])
        self._style_simple_table(ws, len(self.decisions) + 1, len(headers), "TabelaKontrolaCN")
        widths = [10, 48, 48, 18, 16, 14, 12, 28, 78, 70]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
        for row in range(2, len(self.decisions) + 2):
            status = ws.cell(row=row, column=6).value
            if status == STATUS_MISSING:
                for col in range(1, 11):
                    ws.cell(row=row, column=col).fill = FILL_RED
            elif status == STATUS_UNCERTAIN:
                for col in range(1, 11):
                    ws.cell(row=row, column=col).fill = FILL_YELLOW

    def _write_settings_sheet(self, wb: Workbook, declaration_attrs: Dict[str, str], output_path: Path) -> None:
        ws = wb.create_sheet("Ustawienia")
        rows = [
            ["Nazwa", "Wartość"],
            ["Wygenerowano", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Plik wynikowy", str(output_path)],
            ["Liczba pozycji", len(self.decisions)],
            ["CN próg pewny", self.config.get("cn_confident_threshold")],
            ["CN próg niepewny", self.config.get("cn_uncertain_threshold")],
            ["Domyślne warunki dostawy", self.config.get("default_delivery_terms", "")],
            ["Domyślny rodzaj transakcji", self.config.get("default_transaction_type", "")],
            ["Domyślny rodzaj transportu", self.config.get("default_transport_type", "")],
            ["Tryb wartości statystycznej", self.config.get("statistical_value_mode", "")],
            ["Słowników wczytanych", len(self.dicts)],
            ["Kodów taryfy wczytanych", len(self.tariff_entries)],
            ["Pozycji z pustym CN", sum(1 for d in self.decisions if d.get("Status") == STATUS_MISSING)],
            ["Pozycji z niepewnym CN", sum(1 for d in self.decisions if d.get("Status") == STATUS_UNCERTAIN)],
        ]
        for k, v in declaration_attrs.items():
            rows.append([f"Deklaracja.{k}", v])
        if self.warnings:
            rows.append(["", ""])
            rows.append(["OSTRZEŻENIA", ""])
            for w in self.warnings:
                rows.append(["", w])
        for r in rows:
            ws.append(r)
        self._style_simple_table(ws, len(rows), 2, "TabelaUstawienia")
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 96

    def _write_dictionary_sheets(self, wb: Workbook) -> None:
        for code in sorted(self.dicts.keys()):
            d = self.dicts[code]
            ws = wb.create_sheet(f"SLW_{code}")
            ws.append(["Kod", "Opis", "OpisEN", "WaznyOd", "WaznyDo", "Nazwa słownika", "Plik"])
            for row in d.rows:
                ws.append([
                    row.get("Kod", ""), row.get("Opis", ""), row.get("OpisEN", ""), row.get("WaznyOd", ""), row.get("WaznyDo", ""), d.name, d.path.name
                ])
            self._style_simple_table(ws, len(d.rows) + 1, 7, f"TabelaSLW{code}")
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 72
            ws.column_dimensions["C"].width = 56
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 14
            ws.column_dimensions["F"].width = 44
            ws.column_dimensions["G"].width = 32
            ws.freeze_panes = "A2"
            if bool(self.config.get("hide_dictionary_sheets", False)):
                ws.sheet_state = "hidden"

    def _write_tariff_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Taryfa_CN")
        ws.append(["Kod CN", "Kod z odstępami", "Opis", "Ścieżka taryfy"])
        rows = [e for e in self.tariff_entries if e.code.startswith("94") or e.code.startswith("99")]
        for e in rows:
            ws.append([e.code, e.spaced_code, e.description, e.path_text])
        self._style_simple_table(ws, len(rows) + 1, 4, "TabelaTaryfaCN")
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 58
        ws.column_dimensions["D"].width = 100
        ws.freeze_panes = "A2"

    def _style_simple_table(self, ws: Any, rows: int, cols: int, table_name: str) -> None:
        for cell in ws[1]:
            cell.fill = FILL_SUBHEADER
            cell.font = FONT_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER_THIN
        for row in ws.iter_rows(min_row=2, max_row=max(rows, 2), min_col=1, max_col=cols):
            for cell in row:
                cell.border = BORDER_THIN
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.column in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
                    cell.number_format = "@"
        if rows >= 2:
            try:
                ref = f"A1:{get_column_letter(cols)}{rows}"
                table = Table(displayName=table_name, ref=ref)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                ws.add_table(table)
            except Exception:
                pass

    def _apply_validations(self, ws: Any, max_row: int, wb: Workbook) -> None:
        # Formuły listy zapisujemy bez znaku '='. Taki zapis jest bezpieczniejszy dla Excela.
        self._add_list_validation(ws, f"B2:B{max_row}", self._range_for_dict("049") or self._range_for_dict("007"), False,
                                  "Kod kraju UE ze słownika XML 049/007.")
        self._add_list_validation(ws, f"C2:C{max_row}", self._range_for_dict("002"), True,
                                  "Warunki dostawy. Pole może być puste.")
        self._add_list_validation(ws, f"D2:D{max_row}", self._range_for_dict("004"), False,
                                  "Rodzaj transakcji ze słownika XML 004.")
        tariff_last = max(2, len([e for e in self.tariff_entries if e.code.startswith("94") or e.code.startswith("99")]) + 1)
        if "Taryfa_CN" in wb.sheetnames and tariff_last >= 2:
            self._add_list_validation(ws, f"E2:E{max_row}", f"{quote_sheetname('Taryfa_CN')}!$A$2:$A${tariff_last}", True,
                                      "Kod CN z arkusza Taryfa_CN. Czerwone puste pola wymagają ręcznego wyboru.")
        self._add_list_validation(ws, f"F2:F{max_row}", self._range_for_dict("005"), True,
                                  "Rodzaj transportu. Pole może być puste.")
        self._add_list_validation(ws, f"G2:G{max_row}", self._range_for_dict("007"), True,
                                  "Kod kraju pochodzenia ze słownika XML 007.")

    def _range_for_dict(self, code: str) -> str:
        d = self.dicts.get(code)
        if not d or not d.rows:
            return ""
        sheet_name = f"SLW_{code}"
        last = len(d.rows) + 1
        return f"{quote_sheetname(sheet_name)}!$A$2:$A${last}"

    def _add_list_validation(self, ws: Any, cell_range: str, formula_range: str, allow_blank: bool, prompt: str) -> None:
        if not formula_range:
            return
        dv = DataValidation(type="list", formula1=formula_range, allow_blank=allow_blank)
        dv.error = "Wartość spoza słownika. Wybierz kod z listy albo popraw słownik."
        dv.errorTitle = "Błędna wartość"
        dv.prompt = prompt
        dv.promptTitle = "Słownik"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        # showDropDown=False oznacza w OOXML, że strzałka listy ma być widoczna w Excelu.
        dv.showDropDown = False
        ws.add_data_validation(dv)
        dv.add(cell_range)

    def _finalize_workbook(self, wb: Workbook) -> None:
        wb.properties.creator = "Generator INTRASTAT XLSX v2"
        wb.properties.title = "INTRASTAT import XLSX"
        wb.properties.subject = "Wygenerowany plik importu ist@t2"
        wb.active = 0


class GeneratorService:
    def __init__(self, base_dir: Optional[Path] = None):
        self.base_dir = base_dir or get_app_dir()
        self.config_path = self.base_dir / CONFIG_FILE
        self.config = load_json(self.config_path, DEFAULT_CONFIG)
        ensure_dirs(self.base_dir, self.config)

    def save_config(self) -> None:
        save_json(self.config_path, self.config)

    def guess_tariff_path(self) -> str:
        configured = norm_text(self.config.get("tariff_path", ""))
        if configured and Path(configured).exists():
            return configured
        for name in ["taryfa.txt", "taryfa(1).txt"]:
            p = self.base_dir / name
            if p.exists():
                self.config["tariff_path"] = str(p)
                self.save_config()
                return str(p)
        return configured

    def load_current_dicts(self) -> Dict[str, DictionaryData]:
        paths = ensure_dirs(self.base_dir, self.config)
        return DictionaryLoader(self.base_dir, paths["dict"]).load()

    def dict_codes_for_gui(self, code: str) -> List[str]:
        try:
            dicts = self.load_current_dicts()
            d = dicts.get(code)
            return [""] + (d.codes() if d else [])
        except Exception:
            return [""]

    def generate(self, input_xml: Path, tariff_path: Path, progress: Optional[Callable[[int, str], None]] = None) -> Tuple[Path, Dict[str, Any]]:
        t0 = time.perf_counter()
        if progress:
            progress(3, "Sprawdzanie plików...")
        if not input_xml.exists():
            raise FileNotFoundError(f"Nie istnieje plik XML: {input_xml}")
        if not tariff_path.exists():
            raise FileNotFoundError(f"Nie istnieje plik taryfy: {tariff_path}")

        self.config["tariff_path"] = str(tariff_path)
        self.save_config()
        paths = ensure_dirs(self.base_dir, self.config)

        if progress:
            progress(10, "Wczytywanie słowników XML...")
        dicts = DictionaryLoader(self.base_dir, paths["dict"]).load()
        if not dicts:
            raise RuntimeError(f"Nie znaleziono słowników XML. Włóż pliki slownik*.xml do: {paths['dict']}")

        if progress:
            progress(20, "Wczytywanie taryfy CN...")
        tariff_entries = TariffLoader(tariff_path).load()
        if not tariff_entries:
            raise RuntimeError("Nie udało się wczytać kodów CN z taryfa.txt")

        if progress:
            progress(32, "Wczytywanie deklaracji XML...")
        parser = IntrastatXmlParser(input_xml, self.config, dicts)
        items = parser.parse()
        if not items:
            raise RuntimeError("Nie znaleziono żadnych pozycji <Towar> w XML.")

        if progress:
            progress(40, "Dobieranie kodów CN...")
        resolver = CnResolver(
            tariff_entries=tariff_entries,
            confident_threshold=float(self.config.get("cn_confident_threshold", 90.0)),
            uncertain_threshold=float(self.config.get("cn_uncertain_threshold", 80.0)),
        )
        builder = WorkbookBuilder(dicts, tariff_entries, resolver, self.config)

        declaration_no = parser.declaration_attrs.get("NrWlasny", "INTRASTAT") or "INTRASTAT"
        month = parser.declaration_attrs.get("Miesiac", "")
        year = parser.declaration_attrs.get("Rok", "")
        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", f"{declaration_no}_{year}_{month}_{now_stamp()}").strip("_")
        xlsx_path = paths["output"] / f"{safe_name}.xlsx"

        builder.build(items, parser.declaration_attrs, xlsx_path, progress)
        elapsed = time.perf_counter() - t0
        xlsx_path = max(paths["output"].glob(f"{safe_name}*.xlsx"), key=lambda p: p.stat().st_mtime)
        summary = {
            "items_count": len(items),
            "dicts_count": len(dicts),
            "dict_codes": sorted(dicts.keys()),
            "tariff_entries_count": len(tariff_entries),
            "elapsed_seconds": elapsed,
            "warnings_count": len(builder.warnings),
            "missing_cn_count": sum(1 for d in builder.decisions if d.get("Status") == STATUS_MISSING),
            "uncertain_cn_count": sum(1 for d in builder.decisions if d.get("Status") == STATUS_UNCERTAIN),
            "output_dir": str(paths["output"]),
        }
        if progress:
            progress(100, f"Gotowe. Czas: {elapsed:.2f} s")
        return xlsx_path, summary


class App:
    def __init__(self):
        if tk is None:
            raise RuntimeError("Tkinter nie jest dostępny w tym Pythonie.")
        self.service = GeneratorService()
        base_cls = TkinterDnD.Tk if HAS_DND else tk.Tk
        self.root = base_cls()
        self.root.title(APP_NAME)
        self.root.geometry("1020x700")
        self.root.minsize(900, 620)

        self.msg_queue: "queue.Queue[Tuple[str, Any]]" = queue.Queue()
        self.xml_var = tk.StringVar()
        self.tariff_var = tk.StringVar(value=self.service.guess_tariff_path())
        self.dict_dir_var = tk.StringVar(value=str(resolve_path(self.service.config.get("dict_dir", DICT_DIR_NAME), self.service.base_dir)))
        self.delivery_var = tk.StringVar(value=self.service.config.get("default_delivery_terms", ""))
        self.transaction_var = tk.StringVar(value=self.service.config.get("default_transaction_type", "11"))
        self.transport_var = tk.StringVar(value=self.service.config.get("default_transport_type", ""))
        self.stat_mode_var = tk.StringVar(value=self.service.config.get("statistical_value_mode", "blank"))
        self.open_folder_var = tk.BooleanVar(value=bool(self.service.config.get("auto_open_output_folder", False)))
        self.hide_dict_var = tk.BooleanVar(value=bool(self.service.config.get("hide_dictionary_sheets", False)))
        self.confident_var = tk.StringVar(value=str(self.service.config.get("cn_confident_threshold", 90.0)))
        self.uncertain_var = tk.StringVar(value=str(self.service.config.get("cn_uncertain_threshold", 80.0)))
        self.progress_var = tk.IntVar(value=0)
        self.status_var = tk.StringVar(value="Gotowy")
        self._build_ui()
        self.root.after(120, self._poll_queue)

    def _build_ui(self) -> None:
        pad = {"padx": 10, "pady": 6}
        frm = ttk.Frame(self.root, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="Generator XLSX INTRASTAT v2", font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 10))
        info = "Przeciągnij XML deklaracji/Subiekta do pierwszego pola. Taryfa jest zapamiętywana w config.json. Słowniki XML są wczytywane z folderu slowniki i folderu programu."
        if not HAS_DND:
            info += "  Uwaga: przeciąganie nie działa bez tkinterdnd2 — użyj przycisków Wybierz."
        ttk.Label(frm, text=info, wraplength=960).pack(anchor="w", pady=(0, 8))

        fields = ttk.LabelFrame(frm, text="Pliki")
        fields.pack(fill="x", pady=(0, 10))
        self._file_row(fields, "XML deklaracji/Subiekta", self.xml_var, self._select_xml, 0)
        self._file_row(fields, "Taryfa CN / taryfa.txt", self.tariff_var, self._select_tariff, 1)
        self._dir_row(fields, "Folder słowników XML", self.dict_dir_var, self._select_dict_dir, 2)

        options = ttk.LabelFrame(frm, text="Ustawienia domyślne dla brakujących danych")
        options.pack(fill="x", pady=(0, 10))
        for col in range(6):
            options.columnconfigure(col, weight=1)

        delivery_values = self.service.dict_codes_for_gui("002")
        transaction_values = self.service.dict_codes_for_gui("004")
        transport_values = self.service.dict_codes_for_gui("005")

        ttk.Label(options, text="Warunki dostawy").grid(row=0, column=0, sticky="w", **pad)
        self.delivery_combo = ttk.Combobox(options, textvariable=self.delivery_var, values=delivery_values, state="normal", width=12)
        self.delivery_combo.grid(row=0, column=1, sticky="ew", **pad)

        ttk.Label(options, text="Rodzaj transakcji").grid(row=0, column=2, sticky="w", **pad)
        self.transaction_combo = ttk.Combobox(options, textvariable=self.transaction_var, values=transaction_values, state="normal", width=12)
        self.transaction_combo.grid(row=0, column=3, sticky="ew", **pad)

        ttk.Label(options, text="Rodzaj transportu").grid(row=0, column=4, sticky="w", **pad)
        self.transport_combo = ttk.Combobox(options, textvariable=self.transport_var, values=transport_values, state="normal", width=12)
        self.transport_combo.grid(row=0, column=5, sticky="ew", **pad)

        ttk.Label(options, text="Wartość statystyczna").grid(row=1, column=0, sticky="w", **pad)
        ttk.Combobox(options, textvariable=self.stat_mode_var, state="readonly", values=["blank", "copy_invoice_when_required", "copy_invoice_always"]).grid(row=1, column=1, columnspan=2, sticky="ew", **pad)

        ttk.Label(options, text="CN pewny od %").grid(row=1, column=3, sticky="w", **pad)
        ttk.Entry(options, textvariable=self.confident_var, width=10).grid(row=1, column=4, sticky="ew", **pad)
        ttk.Entry(options, textvariable=self.uncertain_var, width=10).grid(row=1, column=5, sticky="ew", **pad)

        checks = ttk.Frame(frm)
        checks.pack(fill="x", pady=(0, 10))
        ttk.Checkbutton(checks, text="Otwórz folder po generowaniu", variable=self.open_folder_var).pack(side="left", padx=(0, 18))
        ttk.Checkbutton(checks, text="Ukryj arkusze słownikowe w XLSX", variable=self.hide_dict_var).pack(side="left")
        ttk.Label(checks, text="  Drugi próg: CN niepewny od %, poniżej tego puste/czerwone.").pack(side="left", padx=(12, 0))

        actions = ttk.Frame(frm)
        actions.pack(fill="x", pady=(0, 10))
        self.generate_btn = ttk.Button(actions, text="Generuj XLSX", command=self._generate_clicked)
        self.generate_btn.pack(side="left", padx=(0, 10))
        ttk.Button(actions, text="Odśwież listy słowników", command=self._refresh_dictionary_combos).pack(side="left", padx=(0, 10))
        ttk.Button(actions, text="Otwórz folder programu", command=lambda: self._open_folder(self.service.base_dir)).pack(side="left", padx=(0, 10))
        ttk.Button(actions, text="Otwórz folder wyników", command=lambda: self._open_folder(resolve_path(self.service.config.get("output_dir", OUTPUT_DIR_NAME), self.service.base_dir))).pack(side="left")

        progress_frame = ttk.Frame(frm)
        progress_frame.pack(fill="x", pady=(0, 10))
        ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100).pack(fill="x", side="left", expand=True, padx=(0, 10))
        ttk.Label(progress_frame, textvariable=self.status_var, width=52).pack(side="right")

        log_frame = ttk.LabelFrame(frm, text="Log")
        log_frame.pack(fill="both", expand=True)
        self.log_text = tk.Text(log_frame, height=14, wrap="word")
        self.log_text.pack(fill="both", expand=True, padx=6, pady=6)
        self.log("Program uruchomiony.")
        self.log(f"Folder programu: {self.service.base_dir}")
        self.log(f"Folder słowników: {self.dict_dir_var.get()}")
        self.log(f"Drag & drop: {'TAK' if HAS_DND else 'NIE - zainstaluj tkinterdnd2'}")
        self.log(f"Fuzzy RapidFuzz: {'TAK' if HAS_RAPIDFUZZ else 'NIE - używam difflib'}")

    def _file_row(self, parent: ttk.LabelFrame, label: str, var: tk.StringVar, command: Callable[[], None], row: int) -> None:
        parent.columnconfigure(1, weight=1)
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=10, pady=8)
        ent = ttk.Entry(parent, textvariable=var)
        ent.grid(row=row, column=1, sticky="ew", padx=10, pady=8)
        ttk.Button(parent, text="Wybierz", command=command).grid(row=row, column=2, padx=10, pady=8)
        self._enable_drop(ent, var)

    def _dir_row(self, parent: ttk.LabelFrame, label: str, var: tk.StringVar, command: Callable[[], None], row: int) -> None:
        parent.columnconfigure(1, weight=1)
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=10, pady=8)
        ent = ttk.Entry(parent, textvariable=var)
        ent.grid(row=row, column=1, sticky="ew", padx=10, pady=8)
        ttk.Button(parent, text="Wybierz", command=command).grid(row=row, column=2, padx=10, pady=8)
        self._enable_drop(ent, var)

    def _enable_drop(self, widget: Any, var: tk.StringVar) -> None:
        if not HAS_DND:
            return
        try:
            widget.drop_target_register(DND_FILES)
            widget.dnd_bind("<<Drop>>", lambda event: var.set(path_from_drop_data(self.root, event.data)))
        except Exception:
            pass

    def _select_xml(self) -> None:
        p = filedialog.askopenfilename(title="Wybierz XML deklaracji", filetypes=[("XML", "*.xml"), ("Wszystkie pliki", "*.*")])
        if p:
            self.xml_var.set(p)

    def _select_tariff(self) -> None:
        p = filedialog.askopenfilename(title="Wybierz taryfa.txt", filetypes=[("TXT", "*.txt"), ("Wszystkie pliki", "*.*")])
        if p:
            self.tariff_var.set(p)

    def _select_dict_dir(self) -> None:
        p = filedialog.askdirectory(title="Wybierz folder słowników XML")
        if p:
            self.dict_dir_var.set(p)
            self._save_options_to_config()
            self._refresh_dictionary_combos()

    def _refresh_dictionary_combos(self) -> None:
        self._save_options_to_config()
        self.delivery_combo.configure(values=self.service.dict_codes_for_gui("002"))
        self.transaction_combo.configure(values=self.service.dict_codes_for_gui("004"))
        self.transport_combo.configure(values=self.service.dict_codes_for_gui("005"))
        self.log("Odświeżono listy domyślnych wartości ze słowników XML.")

    def _save_options_to_config(self) -> None:
        self.service.config["tariff_path"] = self.tariff_var.get().strip()
        self.service.config["dict_dir"] = self.dict_dir_var.get().strip() or DICT_DIR_NAME
        self.service.config["default_delivery_terms"] = self.delivery_var.get().strip().upper()
        self.service.config["default_transaction_type"] = self.transaction_var.get().strip()
        self.service.config["default_transport_type"] = self.transport_var.get().strip()
        self.service.config["statistical_value_mode"] = self.stat_mode_var.get().strip()
        self.service.config["auto_open_output_folder"] = bool(self.open_folder_var.get())
        self.service.config["hide_dictionary_sheets"] = bool(self.hide_dict_var.get())
        try:
            self.service.config["cn_confident_threshold"] = float(str(self.confident_var.get()).replace(",", "."))
            self.service.config["cn_uncertain_threshold"] = float(str(self.uncertain_var.get()).replace(",", "."))
        except Exception:
            self.service.config["cn_confident_threshold"] = 90.0
            self.service.config["cn_uncertain_threshold"] = 80.0
        self.service.save_config()

    def _generate_clicked(self) -> None:
        input_xml = Path(self.xml_var.get().strip().strip('"'))
        tariff = Path(self.tariff_var.get().strip().strip('"'))
        if not str(input_xml) or not input_xml.exists():
            messagebox.showwarning(APP_NAME, "Wskaż poprawny plik XML deklaracji.")
            return
        if not str(tariff) or not tariff.exists():
            messagebox.showwarning(APP_NAME, "Wskaż poprawny plik taryfa.txt.")
            return
        self._save_options_to_config()
        self.generate_btn.config(state="disabled")
        self.progress_var.set(0)
        self.log("Start generowania...")
        t = threading.Thread(target=self._generate_worker, args=(input_xml, tariff), daemon=True)
        t.start()

    def _generate_worker(self, input_xml: Path, tariff: Path) -> None:
        try:
            self.service.config = load_json(self.service.config_path, DEFAULT_CONFIG)
            xlsx_path, summary = self.service.generate(input_xml, tariff, progress=self._progress_from_thread)
            self.msg_queue.put(("done", (xlsx_path, summary)))
        except Exception as exc:
            log_path = log_exception(self.service.base_dir, exc)
            self.msg_queue.put(("error", (str(exc), log_path)))

    def _progress_from_thread(self, percent: int, message: str) -> None:
        self.msg_queue.put(("progress", (percent, message)))

    def _poll_queue(self) -> None:
        try:
            while True:
                kind, payload = self.msg_queue.get_nowait()
                if kind == "progress":
                    percent, message = payload
                    self.progress_var.set(percent)
                    self.status_var.set(message)
                    self.log(message)
                elif kind == "done":
                    xlsx_path, summary = payload
                    self.generate_btn.config(state="normal")
                    self.progress_var.set(100)
                    self.status_var.set("Gotowe")
                    self.log(f"Wygenerowano XLSX: {xlsx_path}")
                    self.log(f"Pozycji: {summary['items_count']}; słowników: {summary['dicts_count']}; taryfa pozycji: {summary['tariff_entries_count']}; czas: {summary['elapsed_seconds']:.2f} s")
                    self.log(f"CN puste/czerwone: {summary['missing_cn_count']}; CN żółte/niepewne: {summary['uncertain_cn_count']}")
                    messagebox.showinfo(APP_NAME, f"Gotowe.\n\nXLSX:\n{xlsx_path}\n\nCzas: {summary['elapsed_seconds']:.2f} s\nCN puste/czerwone: {summary['missing_cn_count']}\nCN żółte/niepewne: {summary['uncertain_cn_count']}")
                    if self.service.config.get("auto_open_output_folder"):
                        self._open_folder(Path(summary["output_dir"]))
                elif kind == "error":
                    self.generate_btn.config(state="normal")
                    msg, log_path = payload
                    self.status_var.set("Błąd")
                    self.log(f"BŁĄD: {msg}")
                    self.log(f"Log błędu: {log_path}")
                    messagebox.showerror(APP_NAME, f"Błąd generowania:\n{msg}\n\nSzczegóły zapisano w:\n{log_path}")
        except queue.Empty:
            pass
        self.root.after(120, self._poll_queue)

    def log(self, message: str) -> None:
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{ts}] {message}\n")
        self.log_text.see("end")

    def _open_folder(self, path: Path) -> None:
        path.mkdir(parents=True, exist_ok=True)
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(path))  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                import subprocess
                subprocess.Popen(["open", str(path)])
            else:
                import subprocess
                subprocess.Popen(["xdg-open", str(path)])
        except Exception as exc:
            self.log(f"Nie udało się otworzyć folderu: {exc}")

    def run(self) -> None:
        self.root.mainloop()


def main() -> int:
    parser = argparse.ArgumentParser(description="Generator XLSX INTRASTAT v2")
    parser.add_argument("--input", help="Ścieżka do XML deklaracji")
    parser.add_argument("--tariff", help="Ścieżka do taryfa.txt")
    parser.add_argument("--dict-dir", help="Folder ze słownikami XML slownik*.xml")
    parser.add_argument("--output-dir", help="Folder wynikowy")
    parser.add_argument("--no-gui", action="store_true", help="Uruchom bez GUI")
    args = parser.parse_args()

    service = GeneratorService()
    if args.dict_dir:
        service.config["dict_dir"] = args.dict_dir
    if args.output_dir:
        service.config["output_dir"] = args.output_dir
    if args.tariff:
        service.config["tariff_path"] = args.tariff
    service.save_config()

    if args.no_gui:
        if not args.input:
            print("Brak --input", file=sys.stderr)
            return 2
        tariff = Path(args.tariff or service.guess_tariff_path())

        def progress(percent: int, msg: str) -> None:
            print(f"{percent:3d}% {msg}")

        try:
            xlsx, summary = service.generate(Path(args.input), tariff, progress=progress)
        except Exception as exc:
            log_path = log_exception(service.base_dir, exc)
            print(f"Błąd: {exc}\nLog: {log_path}", file=sys.stderr)
            return 1
        print("\nGotowe:")
        print(f"XLSX: {xlsx}")
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    app = App()
    app.run()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
